##########################################################################
## Configuration generator from excel workbook
## ver 1.1
##########################################################################
## Features:
## -excel xlsx format is supported
## -unicode titles and cell values in excel workbook are supported, cell
##  value is converted to str type while writing to file 
## -fields in worksheet and template are case insesensitive (converted 
##  to lower case before comparison)
## -configuration directory is cleared before main function (or created
##  if it doesn't exist)
## -several checks are supported:
##  -excel workbook (critical error, stop)
##  -<name> field in worksheet (error, next worksheet)
##  -template file (warning if no additional templates, next worksheet)
##  -empty value for field (warning, next field)
##  -template field not found in worksheet fields (warning, next field)
##  -additional template file (warning, next template)
##  -template file (info if additional templates)
##  -<name> value in line (info, next line)
## -additional templates separated by "template_delim" are supported
## -subset of all worksheets is supported via "wb_set" (by default all
##  worksheets are checked) 
##########################################################################
## Added in 1.1:
## -config filename for template could be provided via "filename_field"
##  in special "Templates" worksheet
##
## "Templates" worksheet example format:
## <name> <filename_field>
## p2p_1  host_1
## p2p_2  host_2
##########################################################################

from openpyxl import load_workbook

import re

import os

##########################################################################
## Basic variables
##
##########################################################################
## path - path to base directories
## config_path - path where configuration files should be created
## template_path - path where template files are located
## template_delim - delimeter for additiional templates
## wb_path - path to excel workbook
## wb_name - excel workbook name
## ws_set - list of excel worksheet name(s)
##########################################################################
path = '.'
config_path = os.path.join(path, 'config', '')
source_config_path = os.path.join(path, 'source', '')
template_path = os.path.join(path, 'template', '')
template_delim = ','
wb_path = os.path.join(path, 'data', '')
wb_name = 'LLD_UR.xlsx'
#wb_name = 'LLD_NBSK_VPN.xlsx'
ws_set = []
##########################################################################
## Main function
##########################################################################
def generate():
	config_dir()
	##########################################################################
	## Try open workbook provided
	##########################################################################
	try:
		wb = load_workbook(wb_path + wb_name, data_only=True)
	except IOError:
		print u'Critical error! Couldn\'t open workbook {}'.format(wb_name)
		return

	if not ws_set:
		for wsi in wb.worksheets:
			ws_set.append(wsi.title)
		
	##########################################################################
	## Check whether "Templates" worksheet is provided and parse it
	##########################################################################
	try:
		ws_templates = wb.get_sheet_by_name('Templates')
		device_name_template = {}
		fields = None
		fields = {}
		for i in range(1, ws_templates.max_column + 1):
			fields['<' + unicode(ws_templates.cell(None, 1, i).value).lower() + '>'] = i
		if u'<name>' in fields:
			name_col = fields[u'<name>']
		else:
			name_col = None
		if u'<filename_field>' in fields:
			filename_field_col = fields[u'<filename_field>']
		else:
			filename_field_col = None
		if name_col and filename_field_col:
			for i in range(2, ws_templates.max_row + 1):
				template_name = ws_templates.cell(None, i, name_col).value
				if not template_name:
					print u'Info! No name provided in line {} in worksheet \"Templates\". Proceeding to next line'.format(i)
					continue
				device_name_template[template_name.lower()] = []
				filename_field = ws_templates.cell(None, i, filename_field_col).value
				if not filename_field:
					device_name_template[template_name] = u'<name>'
				else:
					device_name_template[template_name] = '<' + filename_field.lower() + '>'
	except KeyError:
		ws_templates = None
		
	for wsi in ws_set:
		ws = wb.get_sheet_by_name(wsi)
		ws_name = ws.title
		
		template_name = [ws_name]
		
		print u'\nWorking with worksheet \"{}\"\n'.format(ws_name)
		
		fields = None
		fields = {}
		
		for i in range(1, ws.max_column + 1):
			fields['<' + unicode(ws.cell(None, 1, i).value).lower() + '>'] = i
		
		##########################################################################
		## Check whether <name> field is provided
		##########################################################################
		if u'<name>' in fields:
			name_col = fields[u'<name>']
		elif ws_templates:
			name_col = -1
		else:
			print u'Info! No <name> field is provided in worksheet \"{}\" and "Templates" worksheet was not found. Proceeding to next worksheet'.format(ws_name)
			continue
		
		##########################################################################
		## Check whether additional templates support field is required
		##########################################################################
		if u'<template>' in fields:
			template_col = fields[u'<template>']
			add_template_support = True
		else:
			cur_template_name = template_name
			add_template_support = False
		
		##########################################################################
		## Check whether template file exists
		##########################################################################
		if not os.path.isfile(template_path + template_name[0] + '.conf'):
			if add_template_support:
				print u'Info! Can\'t open template \"{}.conf\". Additional template(s) should be provided for every line'.format(template_name[0])
			else:
				print u'Error! Can\'t open template \"{}.conf\" and additional template support was not detected. Proceeding to next worksheet'.format(template_name[0])
				continue
	
		for i in range(2, ws.max_row + 1):
			if name_col != -1:
				device_name = ws.cell(None, i, name_col).value
				if not device_name and not ws_templates:
					print u'Info! No name provided in line {} in worksheet \"{}\" and "Templates" worksheet was not found. Proceeding to next line'.format(i, ws_name)
					continue
			
			if add_template_support:
				cur_template_name = ws.cell(None, i, template_col).value
				if cur_template_name:
					cur_template_name = cur_template_name.split(template_delim)
				else:
					cur_template_name = template_name
			
			source_conf_file = ''
			
			first = 1
		
			for template in cur_template_name:
				##########################################################################
				## Check whether additional template file exists
				##########################################################################
				if not os.path.isfile(template_path + template + '.conf'):
					print u'Warning! Can\'t open additional template \"{}\" in line {}. Proceeding to next template (or next line)'.format(template, i)
					continue
					
				###
				if re.findall('^sub_', template):
					with open(template_path + template + '.conf', 'r') as template_file:
						sub_line = template_file.read().split('\n')
					if source_conf_file != source_config_path + ws.cell(None, i, fields[u'<source_file_name>']).value:
						try:
							with open(source_config_path + ws.cell(None, i, fields[u'<source_file_name>']).value, 'r') as source_conf_file:
								conf_line = source_conf_file.read()
						except IOError:
							print u'Critical error! Couldn\'t open source file \"{}\" in line {}. Proceeding to next template (or next line)'.format(template, i)
							
					block = 0
					sub_block_exc = 0
					sub_block_var = 0
					sub_block_mod = 0
					sub_block_add = 0
					segment = []
					notes = ws.cell(None, i, fields[u'<notes>']).value
					all_segments = []
					if first:
						all_segments.append([''])
						first = 0
					if notes:
						all_segments.append([''])
						all_segments.append(['##########'])
						all_segments.append([notes])
						all_segments.append(['##########'])
					var = dict()
					for line in sub_line:
						if not line:
							continue
						#print line
						if re.findall('^block', line):
							block = 1
							sub_block_exc = 0
							sub_block_var = 0
							sub_block_mod = 0
							sub_block_add = 0
							continue
						if re.findall('^sub-block-exception', line):
							sub_block_exc = 1
							sub_block_var = 0
							sub_block_mod = 0
							sub_block_add = 0
						if re.findall('^sub-block-var', line):
							sub_block_exc = 0
							sub_block_var = 1
							sub_block_mod = 0
							sub_block_add = 0
							continue
						if re.findall('^sub-block-mod', line):
							sub_block_exc = 0
							sub_block_var = 0
							sub_block_mod = 1
							sub_block_add = 0
							continue
						if re.findall('^sub-block-add', line):
							sub_block_exc = 0
							sub_block_var = 0
							sub_block_mod = 0
							sub_block_add = 1
							continue
						
						if block == 1:
							sub_base = line
							sub_base = re.sub('\[any\]', '.*', sub_base, flags=re.IGNORECASE)
							if re.findall('\[.*\]', sub_base):					
								sub_fields = set(re.findall('\[.*\]', sub_base))
								sub_fields = [sf.lower() for sf in sub_fields]								
								if var:
									for j in sub_fields:
										try:
											for k in var[j]:
												segment_add = re.findall(re.sub(re.sub('\]', '\]', re.sub('\[', '\[', j)), k, sub_base, flags=re.IGNORECASE), conf_line)
												if segment_add:
													segment = segment + ['']
													segment = segment + segment_add													
										except KeyError:
											continue
							else:
								sub_fields = set(re.findall('\<\w+\>', sub_base))
								sub_fields = [sf.lower() for sf in sub_fields]
								for j in sub_fields:
									try:
										if ws.cell(None, i, fields[j]).value is None:
											print u'Warning! Value for field {} was not found in line {} in worksheet \"{}\". Proceeding to next field'.format(j, i, ws_name)
											sub_base = re.sub(j, '', sub_base, flags=re.IGNORECASE)
											continue
										sub_base = re.sub(j, str(ws.cell(None, i, fields[j]).value), sub_base, flags=re.IGNORECASE)
									except KeyError:
										print u'Warning! Field {} was not found in worksheet \"{}\" but provided in template \"{}.conf\". Proceeding to next field'.format(j, ws_name, template)								
								segment_add = re.findall(sub_base, conf_line)
								if segment_add:
									segment = segment + ['']
									segment = segment + segment_add									
							block = 0
							#print segment
							continue

						if sub_block_exc == 1:
							segment_len = len(segment)
							segment_id = 0 
							sub_base = re.sub('\[any\]', '.*', line, flags=re.IGNORECASE)
							while segment_id < segment_len:
								if re.findall(sub_base, segment[segment_id]):									
									segment.remove(re.findall(sub_base, segment[segment_id])[0])
									segment_id = segment_id - 1
									segment_len = segment_len - 1
								segment_id = segment_id + 1
							continue
							
						if sub_block_var == 1:
							sub_base = re.sub('\[any\]', '.*', line, flags=re.IGNORECASE)
							sub_var = re.findall('(\[.*\])', sub_base)[0].lower()
							sub_base = re.sub('\[.*\]', '(.*)', sub_base, flags=re.IGNORECASE)
							var_sub_len = 0
							for j in range(len(segment)):
								if re.findall(sub_base, segment[j]):
									try:
										var[sub_var].append(re.findall(sub_base, segment[j])[0])
									except KeyError:
										var[sub_var] = []
										var[sub_var].append(re.findall(sub_base, segment[j])[0])
									var_sub_len = len(var[sub_var])
								if var_sub_len > 1:
									var[sub_var] = list(set(var[sub_var]))
							continue
						
						if sub_block_mod == 1:
							sub1 = str(ws.cell(None, i, fields[re.findall('\<\w+\>\:\<\w+\>', line)[0].lower().split(':')[0]]).value)
							sub2 = str(ws.cell(None, i, fields[re.findall('\<\w+\>\:\<\w+\>', line)[0].lower().split(':')[1]]).value)
							sub_base = re.sub('\<\w+\>\:\<\w+\>', sub1, line, flags=re.IGNORECASE)
							sub_base = re.sub('\[any\]', '.*', sub_base, flags=re.IGNORECASE)
							for j in range(len(segment)):
								if re.findall(sub_base, segment[j]):
									segment[j] = re.sub(sub1, sub2, segment[j], count=1, flags=re.IGNORECASE)
							continue
							#print segment

						if sub_block_add == 1:
							sub_base = line
							for j in re.findall('\<\w+\>', line):
								sub_base = re.sub(j, str(ws.cell(None, i, fields[j]).value), sub_base, flags=re.IGNORECASE)
							segment.append(sub_base)
							
							continue
							
					all_segments.append(segment)
					#print var
					#for j in all_segments:
					#	for k in j:
					#		print k
					
					if not device_name:
						continue
					
					conf_file = open(config_path + device_name + '.conf', 'a')
					for segment in all_segments:
						for line in segment:
							conf_file.write(line + '\n')
					conf_file.close()				
					continue
				###
				
				with open(template_path + template + '.conf', 'r') as template_file:
					line = template_file.read()
				data_fields = set(re.findall('\<\w+\>', line))
				data_fields = [df.lower() for df in data_fields]
				#print data_fields
				#print fields
				if ws_templates:
					for j in device_name_template:
						if template == j:
							if device_name_template[j] in fields:
								device_name = ws.cell(None, i, fields[device_name_template[j]]).value
								if not device_name:
									print u'Info! No name provided in field {} for template \"{}\" in line {} in worksheet \"{}\". Proceeding to next template/line'.format(device_name_template[j], template, i, ws_name)
								break
							else:
								device_name = None
								print u'Info! Can not found name field {} provided for template \"{}\" in line {} in worksheet \"{}\". Proceeding to next template/line'.format(device_name_template[j], template, i, ws_name)
								break
				if not device_name:
					continue
				conf_file = open(config_path + device_name + '.conf', 'a')
				base_data = line
				if not data_fields:
					print u'Warning! {}: no fields in template {}'.format(device_name, template)
				else:
					for j in data_fields:
						#print j
						try:
							if ws.cell(None, i, fields[j]).value is None:
								print u'Warning! Value for field {} was not found in line {} in worksheet \"{}\". Proceeding to next field'.format(j, i, ws_name)
								base_data = re.sub(j, '', base_data, flags=re.IGNORECASE)
								continue
							#print j
							#print ws.cell(None, i, fields[j]).value
							#base_data = base_data.replace(j, str(ws.cell(None, i, fields[j]).value))
							try:
								base_data = re.sub(j, str(ws.cell(None, i, fields[j]).value), base_data, flags=re.IGNORECASE)
							except Exception as e:
								print u'Critical error! Couldn\'t get cell data'
								print 'Exception:', e
								print 'Line: ', i
								print 'Column:', j
								print 'Cell: ', ws.cell(None, i, fields[j]).value
								return
						except KeyError:
							print u'Warning! Field {} was not found in worksheet \"{}\" but provided in template \"{}.conf\". Proceeding to next field'.format(j, ws_name, template)
				base_data = base_data + '\n'
				conf_file.write(base_data)
				conf_file.close()

##########################################################################
## Config dir utility function
##########################################################################
def config_dir():
	config_dir = config_path[0:len(config_path) - 1]
	if not os.path.exists(config_dir):
		os.makedirs(config_dir)
		print 'Warning! Couldn\'t find config directory. Creating'
	i = 0
	for file in os.listdir(config_dir):
		file_path = os.path.join(config_dir, file)
		if os.path.isfile(file_path):
			os.unlink(file_path)
			i += 1
	print 'Config directory content has been cleared: {} files were deleted'.format(i)
	
generate()

